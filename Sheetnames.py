import openpyxl

wb = openpyxl.load_workbook("MarksList.xlsx")
ws = wb.active
print(ws)
print(wb['Sheet2'])
print(wb.sheetnames) # printing active sheetnames

# creating a new sheet

wb.create_sheet("NEWSHEET")
print(wb.sheetnames)

for sheet in wb:
    print("sheet is: ",sheet)
    print("sheet title is: ", sheet.title)
    
ws3 = wb['Sheet1']
ws3.sheet_properties.tabColor = "1d4be5"

wb.save("MarksList.xlsx")