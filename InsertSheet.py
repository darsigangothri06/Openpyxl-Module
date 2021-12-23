import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

# Create a sheet
ws1 = wb.create_sheet("Master")
# Sheet, Master
# or
ws2 = wb.create_sheet("Master1", 0)  # at first position
# Master1, Sheet, Master
# or
ws3 = wb.create_sheet("Master2", -1)  # create a sheet at penultimate position
# Master1, Sheet, Master2, Master
print(wb.sheetnames)
wb.save("Inserting.xlsx")