# looping through an excel file

import openpyxl
from openpyxl.utils import get_column_letter
wb = openpyxl.load_workbook("MarksList.xlsx")
ws = wb.active  # active sheetname

for row in range(7,15):
    for col in range(1,4):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)
wb.save("MarksList.xlsx")