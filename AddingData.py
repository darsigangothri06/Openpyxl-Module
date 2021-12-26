import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
cnt = 1
for row in range(1,11):
    for col in range(1,11):
        ws.cell(row = row, column = col).value = cnt
        cnt += 1

# Inserting rows
ws.insert_rows(0)
ws.merge_cells("A1:J1")
ws.cell(row = 1, column=1).value = "NUMBERS"

# Displaying data of Excel cells
rowStr = ''
for row in ws.iter_rows(min_row = 1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        rowStr += "{:<5}".format(str(cell.value) + " ")
    rowStr += '\n'
print(rowStr)
wb.save("Numbers.xlsx")