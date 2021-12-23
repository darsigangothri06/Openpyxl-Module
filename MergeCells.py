import openpyxl

wb = openpyxl.load_workbook("MarksList.xlsx")
ws = wb.active

ws.merge_cells("A1:D1")
# unmergse cells
ws.unmerge_cells("A1:D1")  # previous data in other cells is lost
ws.merge_cells("A1:D2")
wb.save("MArksList.xlsx")