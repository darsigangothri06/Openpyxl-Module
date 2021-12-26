import openpyxl

wb = openpyxl.load_workbook("Numbers.xlsx")
ws = wb.active

# Set data into file
def SET(ws):
    cnt = 1
    for row in range(1,11):
        for col in range(1,11):
            ws.cell(row = row, column=col).value = cnt
            cnt += 1
            
# Displaying data
def PRINT(ws):
    rowS = ''
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            rowS += "{:<3}".format(str(cell.value) + " ")
        rowS += '\n'
    print(rowS + '\n')

# Move Range method
ws.move_range("A1:J10", rows=10)

SET(ws)
PRINT(ws)

wb.save("Numbers.xlsx")