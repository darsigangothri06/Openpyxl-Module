import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = openpyxl.Workbook()
# DATA
data = {
    "ONE":{
        "sub1" : 90,
        "sub2" : 89,
        "sub3" : 67,
        "sub4" : 78
    },
    "TWO":{
        "sub1" : 78,
        "sub2": 89,
        "sub3" : 67,
        "sub4" : 65
    }
}

ws = wb.active
ws.title = "GRADES"
headings = ['NUMBER'] + list(data['ONE'].keys())
ws.append(headings)
for person in data:
    grades = [person] + list(data[person].values())
    ws.append(grades)
    
for col in range(2, len('ONE') + 3):
    char = get_column_letter(col)
    ws[char + '4'] = f"=SUM({char + '2'}:{char + '3'})/{len(data)}"
    
# f"=SUM({char + '2'}:{char + '3'})/{len(data)}" CALCULATING SUM

wb.save("AddfromPy.xlsx")