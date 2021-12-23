import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

ws.title = "TITLE"
ws['A1'] = "ai"
ws['B1'] = "b1"

# Appending means adding to the end of worksheet
ws.append(['A1','b1','c1'])
ws.append(['A2','b2','c2'])
ws.append(['end'])

wb.save('NEWEXCEL.xlsx')